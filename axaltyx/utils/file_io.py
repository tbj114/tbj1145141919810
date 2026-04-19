#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
文件读写模块 - 支持 .tbj 文件格式和外部格式
"""

import json
import gzip
import os
import csv
import struct
import chardet
from axaltyx.core.data.dataset import Dataset
from axaltyx.core.data.variable import Variable
from axaltyx.utils.config import ConfigManager
from axaltyx.i18n import I18nManager

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    print("openpyxl not installed, Excel support disabled")
    load_workbook = None
    Workbook = None

try:
    import pyreadstat
except ImportError:
    print("pyreadstat not installed, SPSS support disabled")
    pyreadstat = None


class FileIO:
    """文件读写类"""
    
    @staticmethod
    def save_dataset(dataset, filepath):
        """保存数据集到 .tbj 文件
        
        Args:
            dataset: Dataset 对象
            filepath: 文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            # 转换数据集为字典
            data = dataset.to_dict()
            
            # 为了与文档兼容，我们可以同时支持两种格式，
            # 或者直接使用 .tbj 扩展名的压缩 JSON 格式
            # 按照文档的建议，我们使用 .tbj 格式
            with gzip.open(filepath, 'wt', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            # 更新数据集的文件名和修改状态
            dataset.filename = filepath
            dataset.modified = False
            
            # 添加到最近文件列表
            FileIO.add_recent_file(filepath)
            
            return True
        except Exception as e:
            print(f"保存文件失败: {e}")
            return False
    
    @staticmethod
    def load_dataset(filepath):
        """从 .tbj 文件加载数据集
        
        Args:
            filepath: 文件路径
            
        Returns:
            Dataset: 加载的数据集对象，失败返回 None
        """
        try:
            # 读取压缩的 JSON 文件
            with gzip.open(filepath, 'rt', encoding='utf-8') as f:
                data = json.load(f)
            
            # 从字典创建数据集
            dataset = Dataset.from_dict(data)
            dataset.filename = filepath
            
            # 添加到最近文件列表
            FileIO.add_recent_file(filepath)
            
            return dataset
        except Exception as e:
            print(f"加载文件失败: {e}")
            return None
    
    @staticmethod
    def create_new_dataset():
        """创建新的空数据集
        
        Returns:
            Dataset: 新的数据集对象
        """
        return Dataset(rows=100, cols=100)
    
    @staticmethod
    def add_recent_file(filepath):
        """添加文件到最近文件列表
        
        Args:
            filepath: 文件路径
        """
        config = ConfigManager()
        recent_files = config.get("recent_files", [])
        
        # 移除已存在的相同文件
        recent_files = [f for f in recent_files if f != filepath]
        
        # 添加到列表开头
        recent_files.insert(0, filepath)
        
        # 限制最近文件数量为 10
        recent_files = recent_files[:10]
        
        # 保存回配置
        config.set("recent_files", recent_files)
        config.save()
    
    @staticmethod
    def get_recent_files():
        """获取最近文件列表
        
        Returns:
            list: 最近文件路径列表
        """
        config = ConfigManager()
        return config.get("recent_files", [])
    
    @staticmethod
    def clear_recent_files():
        """清空最近文件列表"""
        config = ConfigManager()
        config.set("recent_files", [])
        config.save()
    
    @staticmethod
    def load_csv(filepath):
        """从 CSV 文件加载数据集
        
        Args:
            filepath: 文件路径
            
        Returns:
            Dataset: 加载的数据集对象，失败返回 None
        """
        try:
            # 自动检测编码
            with open(filepath, 'rb') as f:
                result = chardet.detect(f.read(10000))
                encoding = result['encoding'] or 'utf-8'
            
            # 读取 CSV 文件
            with open(filepath, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            if not rows:
                return None
            
            # 创建数据集
            headers = rows[0]
            data_rows = rows[1:]
            cols = len(headers)
            rows_count = len(data_rows)
            
            dataset = Dataset(rows=rows_count, cols=cols)
            
            # 设置变量名
            for i, header in enumerate(headers):
                if i < len(dataset.variables):
                    dataset.variables[i].name = header
            
            # 填充数据
            for row_idx, row_data in enumerate(data_rows):
                for col_idx, value in enumerate(row_data):
                    if col_idx < cols:
                        # 尝试转换为数值
                        try:
                            if '.' in value:
                                value = float(value)
                            else:
                                value = int(value)
                        except:
                            pass
                        dataset.set_value(row_idx, col_idx, value)
            
            dataset.filename = filepath
            FileIO.add_recent_file(filepath)
            return dataset
        except Exception as e:
            print(f"加载 CSV 文件失败: {e}")
            return None
    
    @staticmethod
    def save_csv(dataset, filepath):
        """将数据集保存为 CSV 文件
        
        Args:
            dataset: Dataset 对象
            filepath: 文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # 写入表头
                headers = [var.name for var in dataset.variables]
                writer.writerow(headers)
                
                # 写入数据
                for row in range(dataset.rows):
                    row_data = []
                    for col in range(dataset.cols):
                        value = dataset.get_value(row, col)
                        row_data.append(value if value is not None else '')
                    writer.writerow(row_data)
            
            dataset.filename = filepath
            dataset.modified = False
            FileIO.add_recent_file(filepath)
            return True
        except Exception as e:
            print(f"保存 CSV 文件失败: {e}")
            return False
    
    @staticmethod
    def load_excel(filepath):
        """从 Excel 文件加载数据集
        
        Args:
            filepath: 文件路径
            
        Returns:
            Dataset: 加载的数据集对象，失败返回 None
        """
        try:
            if load_workbook is None:
                print("openpyxl not installed")
                return None
            
            # 读取 Excel 文件
            wb = load_workbook(filepath)
            ws = wb.active
            
            # 获取数据
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            
            if not rows:
                return None
            
            # 创建数据集
            headers = rows[0]
            data_rows = rows[1:]
            cols = len(headers)
            rows_count = len(data_rows)
            
            dataset = Dataset(rows=rows_count, cols=cols)
            
            # 设置变量名
            for i, header in enumerate(headers):
                if i < len(dataset.variables):
                    dataset.variables[i].name = str(header) if header is not None else f"VAR{i+1:05d}"
            
            # 填充数据
            for row_idx, row_data in enumerate(data_rows):
                for col_idx, value in enumerate(row_data):
                    if col_idx < cols:
                        dataset.set_value(row_idx, col_idx, value)
            
            dataset.filename = filepath
            FileIO.add_recent_file(filepath)
            return dataset
        except Exception as e:
            print(f"加载 Excel 文件失败: {e}")
            return None
    
    @staticmethod
    def save_excel(dataset, filepath):
        """将数据集保存为 Excel 文件
        
        Args:
            dataset: Dataset 对象
            filepath: 文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            if Workbook is None:
                print("openpyxl not installed")
                return False
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            
            # 写入表头
            headers = [var.name for var in dataset.variables]
            ws.append(headers)
            
            # 写入数据
            for row in range(dataset.rows):
                row_data = []
                for col in range(dataset.cols):
                    value = dataset.get_value(row, col)
                    row_data.append(value)
                ws.append(row_data)
            
            # 保存文件
            wb.save(filepath)
            
            dataset.filename = filepath
            dataset.modified = False
            FileIO.add_recent_file(filepath)
            return True
        except Exception as e:
            print(f"保存 Excel 文件失败: {e}")
            return False
    
    @staticmethod
    def load_spss(filepath):
        """从 SPSS .sav 文件加载数据集
        
        Args:
            filepath: 文件路径
            
        Returns:
            Dataset: 加载的数据集对象，失败返回 None
        """
        try:
            if pyreadstat is None:
                print("pyreadstat not installed")
                return None
            
            # 读取 SPSS 文件
            df, meta = pyreadstat.read_sav(filepath)
            
            # 创建数据集
            rows = len(df)
            cols = len(df.columns)
            dataset = Dataset(rows=rows, cols=cols)
            
            # 设置变量属性
            for i, col in enumerate(df.columns):
                if i < len(dataset.variables):
                    var = dataset.variables[i]
                    var.name = col
                    
                    # 设置变量标签
                    if col in meta.column_names_to_labels:
                        var.label = meta.column_names_to_labels[col]
                    
                    # 设置值标签
                    if col in meta.value_labels:
                        var.value_labels = meta.value_labels[col]
                    
                    # 设置变量类型
                    if df[col].dtype == 'object':
                        var.type = Variable.TYPE_STRING
                    else:
                        var.type = Variable.TYPE_NUMERIC
            
            # 填充数据
            for row_idx, row in df.iterrows():
                for col_idx, (col, value) in enumerate(row.items()):
                    if col_idx < cols:
                        dataset.set_value(row_idx, col_idx, value)
            
            dataset.filename = filepath
            FileIO.add_recent_file(filepath)
            return dataset
        except Exception as e:
            print(f"加载 SPSS 文件失败: {e}")
            return None
    
    @staticmethod
    def get_file_filters():
        """获取文件对话框过滤器
        
        Returns:
            str: 文件过滤器字符串
        """
        i18n = I18nManager()
        return i18n.t("app.file_filter")
    
    @staticmethod
    def get_import_filters():
        """获取导入文件过滤器
        
        Returns:
            str: 导入文件过滤器字符串
        """
        i18n = I18nManager()
        return i18n.t("app.file_filter_import")
    
    @staticmethod
    def get_export_filters():
        """获取导出文件过滤器
        
        Returns:
            str: 导出文件过滤器字符串
        """
        i18n = I18nManager()
        return i18n.t("app.file_filter_export")
    
    @staticmethod
    def load_file(filepath):
        """根据文件扩展名自动加载文件
        
        Args:
            filepath: 文件路径
            
        Returns:
            Dataset: 加载的数据集对象，失败返回 None
        """
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext == '.tbj':
            return FileIO.load_dataset(filepath)
        elif ext == '.csv':
            return FileIO.load_csv(filepath)
        elif ext in ['.xlsx', '.xls']:
            return FileIO.load_excel(filepath)
        elif ext == '.sav':
            return FileIO.load_spss(filepath)
        else:
            i18n = I18nManager()
            print(f"{i18n.t('app.invalid_format')}: {ext}")
            return None
    
    @staticmethod
    def save_file(dataset, filepath):
        """根据文件扩展名自动保存文件
        
        Args:
            dataset: Dataset 对象
            filepath: 文件路径
            
        Returns:
            bool: 保存是否成功
        """
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext == '.tbj':
            return FileIO.save_dataset(dataset, filepath)
        elif ext == '.csv':
            return FileIO.save_csv(dataset, filepath)
        elif ext in ['.xlsx', '.xls']:
            return FileIO.save_excel(dataset, filepath)
        else:
            i18n = I18nManager()
            print(f"{i18n.t('app.invalid_format')}: {ext}")
            return False
